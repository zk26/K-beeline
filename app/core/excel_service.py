"""Excel 账号表服务：读取、备注写回、文件锁定检测与处理"""
from __future__ import annotations

import msvcrt
import os
import subprocess
import time
import zipfile
from datetime import datetime
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from app.core.constants import STATUS_CODES_ACTIVATE, STATUS_CODES_POST
from app.core.models import RunMode, StudentRecord

try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False


class ExcelLockedError(Exception):
    """Excel 文件被其他程序占用"""

    def __init__(self, path: str, locked_by: Optional[list] = None):
        self.path = path
        self.locked_by = locked_by or []
        names = ", ".join(f"{p['name']}(PID:{p['pid']})" for p in self.locked_by) or "未知程序"
        super().__init__(f"Excel文件正被占用（{names}），请关闭后重试")


def check_file_locked(file_path: str) -> Tuple[bool, Optional[list]]:
    """检查文件是否被占用。返回 (是否锁定, 占用进程列表)"""
    try:
        with open(file_path, "r+b") as f:
            msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
            msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
        return False, None
    except IOError:
        locked_by = []
        if PSUTIL_AVAILABLE:
            try:
                for proc in psutil.process_iter(["pid", "name"]):
                    try:
                        for file_info in proc.open_files():
                            if os.path.normpath(file_info.path) == os.path.normpath(file_path):
                                locked_by.append({"pid": proc.info["pid"], "name": proc.info["name"]})
                    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                        pass
            except Exception:
                pass
        return True, locked_by
    except Exception:
        return True, None


def force_close_excel_processes(file_path: str, log=print) -> int:
    """强制关闭占用指定 Excel 文件的进程，返回关闭数量"""
    if not PSUTIL_AVAILABLE:
        try:
            for image in ("EXCEL.EXE", "wps.exe", "et.exe"):
                subprocess.run(["taskkill", "/F", "/IM", image], capture_output=True, timeout=5)
            time.sleep(1)
            return 1
        except Exception as e:
            log(f"使用taskkill关闭进程时出错: {str(e)}")
            return 0

    closed_count = 0
    try:
        for proc in psutil.process_iter(["pid", "name"]):
            try:
                for file_info in proc.open_files():
                    if os.path.normpath(file_info.path) == os.path.normpath(file_path):
                        proc_name = proc.info["name"].lower()
                        if "excel" in proc_name or "wps" in proc_name or "et" in proc_name:
                            try:
                                proc.terminate()
                                proc.wait(timeout=3)
                                closed_count += 1
                            except psutil.TimeoutExpired:
                                proc.kill()
                                closed_count += 1
                            except Exception:
                                pass
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass
    except Exception as e:
        log(f"强制关闭进程时出错: {str(e)}")
    return closed_count


def ensure_file_unlocked(file_path: str) -> None:
    """确认文件未被占用，否则抛 ExcelLockedError"""
    is_locked, locked_by = check_file_locked(file_path)
    if is_locked:
        raise ExcelLockedError(file_path, locked_by)


def read_student_data(excel_path: str, log=print):
    """读取账号表。返回 (workbook, worksheet, records, headers)。

    表头需包含：学号、备注
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel文件不存在: {excel_path}")

    ensure_file_unlocked(excel_path)

    try:
        wb = load_workbook(excel_path, read_only=False, data_only=True)
        ws = wb.active

        # 自动检测表头行（找到包含"学号"和"备注"的行）
        header_row = None
        for row in range(1, min(ws.max_row + 1, 20)):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and "学号" in str(cell_val):
                header_row = row
                break
        if header_row is None:
            wb.close()
            raise ValueError("Excel文件中未找到'学号'表头，请使用正确的账号表模板")

        headers = {}
        for col in range(1, min(ws.max_column + 1, 20)):
            header = ws.cell(row=header_row, column=col).value
            if header:
                headers[str(header).strip()] = col
        # 兜底：max_column 可能因为格式蔓延变得很大，实际有数据的列很少

        if "备注" not in headers:
            wb.close()
            raise ValueError("Excel文件中缺少'备注'列，请使用正确的账号表模板")

        records: List[StudentRecord] = []
        for row in range(header_row + 1, ws.max_row + 1):
            student_id = ws.cell(row=row, column=headers["学号"]).value
            remark = ws.cell(row=row, column=headers["备注"]).value
            if student_id:
                sid = str(student_id).strip()
                # 跳过模板示例行（备注含"示例"）
                if remark and "示例" in str(remark).strip():
                    continue
                records.append(StudentRecord(row=row, student_id=sid, remark=remark or ""))

        log(f"成功读取 {len(records)} 条账号记录")
        return wb, ws, records, headers

    except zipfile.BadZipFile as e:
        raise Exception(
            f"Excel文件格式错误或已损坏: {str(e)}\n"
            "可能原因：文件正被其他程序打开 / 文件损坏 / 文件正在写入。\n"
            "请关闭所有打开该文件的程序后重试。"
        )
    except PermissionError as e:
        raise Exception(f"Excel文件被占用，无法读取: {str(e)}\n请关闭 Excel、WPS 等程序后重试。")


def _is_local_ip(ip_address: str) -> bool:
    """判断是否是本地（内网）IP"""
    if not ip_address or ip_address == "未知":
        return True
    try:
        import ipaddress
        ip = ipaddress.ip_address(ip_address)
        return ip.is_private or ip.is_loopback or ip.is_link_local
    except Exception:
        return ip_address.startswith((
            "127.", "192.168.", "10.",
            "172.16.", "172.17.", "172.18.", "172.19.", "172.20.",
            "172.21.", "172.22.", "172.23.", "172.24.", "172.25.",
            "172.26.", "172.27.", "172.28.", "172.29.", "172.30.", "172.31.",
        ))


def build_remark(
    mode: RunMode,
    status_code: int,
    password_reset: bool = False,
    error_message: str = "",
    login_password: str = "",
    login_ip: str = "未知",
    login_location: str = "未知",
    reset_password: str = "A123456a",
) -> str:
    """按模式构建备注文本。

    发帖模式: 序号.发帖状态-是否重置[登录密码]-位置-IP
    激活模式: 序号.激活状态-密码重置状态[登录密码]-时间
    """
    if mode == RunMode.POST:
        status_info = STATUS_CODES_POST.get(status_code, "未知状态")
        reset_status = "已重置" if password_reset else "未重置"

        if password_reset:
            password_display = reset_password
        elif status_code == 1:
            password_display = "初始密码"
        else:
            password_display = login_password if login_password else "未知"

        location_display = login_location if login_location else "未知"
        remark = f"{status_code}.{status_info}-{reset_status}[{password_display}]-{location_display}"
        if not _is_local_ip(login_ip):
            remark += f"-{login_ip if login_ip else '未知'}"
        return remark

    # 激活模式
    status_info = STATUS_CODES_ACTIVATE.get(status_code, "未知状态")
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if status_code == 1:
        reset_status = "密码重置成功" if password_reset else "密码未重置"
        password_display = login_password if login_password else "未知"
        return f"{status_code}.{status_info}-{reset_status}[{password_display}]-{current_time}"
    if error_message:
        return f"{status_code}.{status_info}-{error_message}-{current_time}"
    return f"{status_code}.{status_info}-{current_time}"


def update_remark(ws, row: int, remark_col: int, remark: str, log=print) -> bool:
    """把备注写回工作表（不落盘，由调用方统一 save）"""
    try:
        ws.cell(row=row, column=remark_col, value=remark)
        log(f"更新备注: {remark}")
        return True
    except Exception as e:
        log(f"更新备注失败: {str(e)}")
        return False
